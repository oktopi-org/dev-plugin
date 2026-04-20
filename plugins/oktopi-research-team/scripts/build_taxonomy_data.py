#!/usr/bin/env python3
"""Extract Oktopi taxonomy configuration into plugin data and agent/skill artifacts.

Reads from a local clone of oktopi-org/Taxonomy-config and writes:

    plugins/oktopi-research-team/data/
        functions.json
        stage-gates.json
        modes.json
        heatmap/<modality>.json
        questions/<modality>/<FUNCTION_CODE>.json

    plugins/oktopi-research-team/agents/<function>-reviewer.md      (one per function)
    plugins/oktopi-research-team/skills/stage-gate-sg<N>/SKILL.md   (one per stage gate)

Run from the repo root:
    python3 plugins/oktopi-research-team/scripts/build_taxonomy_data.py \
        --taxonomy ../Taxonomy-config

All data is regenerated from scratch each run; the produced files are the
runtime reference material used by the agents and skills.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

# --------------------------------------------------------------------------
# Static metadata
# --------------------------------------------------------------------------
MODES = {
    "SR": {"name": "Strategic Readiness", "prefix": "C_"},
    "OE": {"name": "Operational Execution", "prefix": "OE_"},
    "DD": {"name": "Due Diligence", "prefix": "DD_"},
    "RS": {"name": "Regulatory Submission", "prefix": "RS_"},
}

STAGE_GATE_GOALS = {
    "SG1": {
        "name": "Initiate Discovery Program & Target Identification",
        "goal": "Validate a biologically credible target, a defensible rationale, and a "
                "screening plan that can yield testable hypotheses.",
        "focus": ["target validation", "MoA rationale", "IP freedom-to-operate",
                   "discovery plan & budget"],
    },
    "SG2": {
        "name": "Lead Nomination / Early Optimization",
        "goal": "Nominate a lead series with acceptable developability signals and a "
                "clear optimization plan to reach clinical candidate criteria.",
        "focus": ["lead criteria met", "DMPK/tox early signals", "scaffold IP",
                   "formulation feasibility"],
    },
    "SG3": {
        "name": "Clinical Candidate Selection / IND-Enabling Entry",
        "goal": "Confirm a single clinical candidate with an approved IND-enabling plan "
                "across CMC, tox, PK/PD, and clinical design.",
        "focus": ["candidate selection data", "IND-enabling study plan",
                   "CMC readiness for GLP/GMP", "target product profile draft"],
    },
    "SG4": {
        "name": "Pre-Clinical to FIH-Enabling (CMC, PK/PD, Tox Ready)",
        "goal": "Complete IND-enabling package: GLP tox complete, CMC GMP drug product "
                "and IND Module 3, FIH protocol and starting dose justification.",
        "focus": ["GLP tox package", "CMC GMP release", "FIH protocol", "starting dose"],
    },
    "SG5": {
        "name": "First-in-Human (FIH) Clinical Program Initiation",
        "goal": "Open IND, activate sites, dose the first subject safely and collect "
                "data suitable for dose-escalation and early PK/PD decisions.",
        "focus": ["IND clearance", "site activation", "FPFV safety",
                   "dose-escalation readiness"],
    },
    "SG6": {
        "name": "Proof of Concept (POC) & Phase 2 Readiness",
        "goal": "Demonstrate clinical proof of concept and confirm Phase 2 design, dose, "
                "population, endpoints, and operational feasibility.",
        "focus": ["POC criteria met", "Phase 2 protocol", "biomarker strategy",
                   "scale-up CMC"],
    },
    "SG7": {
        "name": "Confirmatory & Phase 3 Readiness (Reg & Commercial aligned)",
        "goal": "Lock Phase 3 design with regulatory and commercial alignment; confirm "
                "CMC registration strategy and site/CRO readiness.",
        "focus": ["Phase 3 design", "EoP2 alignment", "registration CMC", "TPP lock"],
    },
    "SG8": {
        "name": "Regulatory Submission / Payer Engagement / Pre-Launch",
        "goal": "Submit a reviewable NDA/BLA, finalize payer value dossier, and prepare "
                "commercial launch readiness.",
        "focus": ["NDA/BLA submission", "labeling strategy", "payer value dossier",
                   "launch readiness"],
    },
    "SG9": {
        "name": "Launch & Post-Market Lifecycle Management",
        "goal": "Execute a compliant launch and sustain post-market obligations, "
                "lifecycle planning, and RWE/safety monitoring.",
        "focus": ["PSUR/PBRER", "REMS/RMP", "LCM indications", "RWE commitments"],
    },
}

# Function metadata — each agent embodies the role goal, not just a question list.
# `role` = seasoned persona, `mission` = outcome they deliver, `mandate` = what they
# own, `principles` = how they reason. Question bank lives in the JSON data files.
FUNCTIONS = [
    {
        "code": "CMC", "biologics_code": "BBCMC",
        "name": "Chemistry, Manufacturing, and Controls",
        "role": "Head of CMC with 20+ years across small-molecule and biologics process development, GMP manufacturing, analytical method validation, and CMC regulatory strategy.",
        "mission": "Prove the molecule can be manufactured reproducibly at the scale and quality the clinical and commercial plans require, with a regulatory package that survives an inspection.",
        "mandate": [
            "Drug substance & drug product development",
            "Analytical methods and specifications",
            "Stability program",
            "GMP manufacturing scale-up and process validation",
            "Supply chain (API, DP, comparator) and release testing",
            "CMC regulatory modules (Module 3, comparability, post-approval changes)",
        ],
        "triggers": [
            "drug substance or drug product development",
            "GMP manufacturing, tech transfer, or process validation",
            "analytical methods, specifications, or stability",
            "CMC regulatory package, Module 3, or comparability",
            "supply chain, API sourcing, or release testing",
            "biologics CMC: cell line, upstream/downstream, aggregation",
        ],
    },
    {
        "code": "PT", "biologics_code": "BBPT",
        "name": "Pharmacology & Toxicology",
        "role": "Head of Nonclinical Safety with deep experience in IND-enabling GLP toxicology, safety pharmacology, ADME, and translational risk assessment.",
        "mission": "Establish that the asset's mechanism and safety profile support human dosing at the proposed starting dose, with justified duration and species coverage for every stage-gate.",
        "mandate": [
            "GLP tox strategy and species selection",
            "Safety pharmacology and genotoxicity",
            "DMPK / ADME characterization",
            "Starting-dose / FIH dose justification (MABEL, HED, PAD)",
            "Juvenile, reproductive, carcinogenicity planning where applicable",
            "Integrated nonclinical package for IND/CTA modules",
        ],
        "triggers": [
            "IND-enabling toxicology, GLP tox, or species selection",
            "safety pharmacology, genotoxicity, or carcinogenicity",
            "DMPK, ADME, or preclinical PK",
            "starting dose, MABEL, HED, or first-in-human justification",
            "juvenile, reproductive, or developmental tox",
            "biologics-specific nonclinical concerns (immunogenicity, CRS, cytokine release)",
        ],
    },
    {
        "code": "TM", "biologics_code": "BBTM",
        "name": "Translational Medicine",
        "role": "VP of Translational Medicine bridging nonclinical and clinical science — biomarkers, PK/PD modeling, target engagement, and proof-of-mechanism design.",
        "mission": "Build the quantitative bridge from preclinical data to clinical proof-of-concept so that Phase 1/2 decisions are data-driven, not hopeful.",
        "mandate": [
            "Biomarker strategy (target engagement, PD, patient selection, safety)",
            "Translational PK/PD models",
            "Proof-of-mechanism and proof-of-concept design",
            "Companion diagnostic strategy (if applicable)",
            "Reverse translation from clinical signals",
            "Dose-prediction and dose-justification support",
        ],
        "triggers": [
            "biomarker strategy, target engagement, or patient selection biomarkers",
            "translational PK/PD modeling, human dose prediction",
            "proof-of-mechanism or proof-of-concept design",
            "companion diagnostic (CDx) strategy",
            "reverse translation from clinical signals",
            "bridging preclinical data to Phase 1/2 decisions",
        ],
    },
    {
        "code": "CP", "biologics_code": "BBCP",
        "name": "Clinical Pharmacology",
        "role": "Clinical Pharmacology Lead responsible for human PK/PD, dose-exposure-response, DDI, special populations, and dose-selection.",
        "mission": "Ensure every dose decision — starting, escalation, Phase 3, and label — is anchored in a defensible exposure-response story.",
        "mandate": [
            "Human PK characterization (linearity, accumulation, variability)",
            "Exposure-response modeling (efficacy + safety)",
            "DDI strategy (in vitro + clinical)",
            "Special populations (renal, hepatic, pediatric, geriatric, pregnancy)",
            "Pop-PK / QSP modeling support",
            "Label dose recommendation and dose-finding design",
        ],
        "triggers": [
            "clinical PK, exposure-response, or dose-finding",
            "DDI strategy or in vitro / clinical DDI studies",
            "special populations (renal, hepatic, pediatric, pregnancy)",
            "pop-PK, PBPK, or QSP modeling",
            "Project Optimus dose optimization",
            "label dose justification",
        ],
    },
    {
        "code": "CDM", "biologics_code": "BBCDM",
        "name": "Clinical Development / Medical",
        "role": "Chief Medical Officer / Clinical Development Lead with hands-on experience in Phase 1–3 design, endpoint selection, investigator networks, and benefit-risk framing.",
        "mission": "Deliver a clinical development plan that produces evidence capable of supporting approval, label, reimbursement, and uptake in the target indication.",
        "mandate": [
            "Target Product Profile (TPP) and clinical development plan (CDP)",
            "Phase 1–3 protocol design (population, endpoints, comparators)",
            "Medical monitoring, safety oversight, DSMB interaction",
            "Benefit-risk assessment",
            "Post-approval study commitments / lifecycle evidence",
            "KOL / investigator engagement on clinical strategy",
        ],
        "triggers": [
            "Target Product Profile (TPP) or clinical development plan (CDP)",
            "Phase 1, 2, or 3 protocol design",
            "endpoint selection, comparator arm, or patient population",
            "benefit-risk assessment or clinical no-go criteria",
            "DSMB charter, medical monitoring, or safety review",
            "indication strategy or lifecycle clinical evidence",
        ],
    },
    {
        "code": "SAF", "biologics_code": "BSAF",
        "name": "Clinical Safety",
        "role": "Head of Pharmacovigilance / Clinical Safety responsible for integrated safety analysis, signal detection, and benefit-risk through development and post-market.",
        "mission": "Detect and characterize safety signals early, keep benefit-risk defensible, and maintain a submission-ready safety narrative at every gate.",
        "mandate": [
            "Safety Management Plan (SMP) / Medical Monitoring Plan",
            "SAE/SUSAR processing and expedited reporting",
            "Integrated safety summary and DSUR/PBRER",
            "Risk Management Plan (RMP) / REMS",
            "Signal detection and benefit-risk updates",
            "Post-market PV surveillance and safety DB readiness",
        ],
        "triggers": [
            "SAE, SUSAR, or expedited safety reporting",
            "DSUR, PBRER, or integrated safety summary",
            "Risk Management Plan (RMP) or REMS",
            "safety signal detection or benefit-risk update",
            "pharmacovigilance, safety database, or safety narrative",
            "biologic-specific safety: CRS, cytokine release, immunogenicity events",
        ],
    },
    {
        "code": "COP", "biologics_code": "BCOP",
        "name": "Clinical Operations",
        "role": "VP of Clinical Operations with end-to-end accountability for site activation, enrollment, vendor oversight, and data readiness across Phase 1–3.",
        "mission": "Execute the clinical plan on time, on budget, and at quality — with every database lock supporting the intended regulatory decision.",
        "mandate": [
            "Site / CRO selection and oversight",
            "Enrollment planning and risk-based monitoring",
            "Drug-supply logistics (IRT, depots, blinding)",
            "Vendor management (central labs, imaging, IRT, eCOA)",
            "Protocol deviations and quality issues",
            "Data cleaning, database lock, TMF inspection-readiness",
        ],
        "triggers": [
            "site activation, feasibility, or country selection",
            "enrollment forecasting or risk-based monitoring",
            "CRO selection, oversight, or governance",
            "central labs, imaging, IRT, or eCOA vendors",
            "TMF, data cleaning, or database lock",
            "drug supply logistics, depots, or blinding integrity",
        ],
    },
    {
        "code": "STAT", "biologics_code": "BSTAT",
        "name": "Biostatistics",
        "role": "Chief Biostatistician aligned with ICH E9(R1) estimand framework, experienced in adaptive designs, multiplicity, missing-data handling, and regulatory submissions.",
        "mission": "Guarantee that the statistical design and analysis actually answer the clinical question with controlled error rates and defensible conclusions at inspection.",
        "mandate": [
            "Trial design (RCT, adaptive, single-arm, external control)",
            "Sample-size / power / estimand specification",
            "Statistical Analysis Plan (SAP), multiplicity, interim analyses",
            "Randomization and blinding integrity",
            "Missing-data strategy and sensitivity analyses",
            "Integrated summary of efficacy / safety (ISE/ISS)",
        ],
        "triggers": [
            "sample size, power, or estimand (ICH E9 R1)",
            "SAP, multiplicity, or alpha-spending",
            "adaptive design, group-sequential, or external control",
            "randomization, blinding, or allocation concealment",
            "missing data, imputation, or sensitivity analyses",
            "ISE/ISS, CSR TLFs, or statistical filing readiness",
        ],
    },
    {
        "code": "REG", "biologics_code": "BBREG",
        "name": "Regulatory Affairs",
        "role": "Head of Global Regulatory Affairs with FDA, EMA, PMDA, and ICH submission experience, including accelerated pathways, orphan designation, and advisory committees.",
        "mission": "Land the regulatory strategy and submissions required to reach a reimbursable label in each priority market on the target timeline.",
        "mandate": [
            "Regulatory strategy (pathway, designations, agency interactions)",
            "Pre-IND / EoP / pre-BLA meeting strategy and briefing docs",
            "IND / NDA / BLA / CTA / MAA submission planning",
            "Labeling strategy and negotiation",
            "Post-approval commitments and variations",
            "Global regulatory alignment and lifecycle maintenance",
        ],
        "triggers": [
            "IND, NDA, BLA, CTA, or MAA submission",
            "FDA, EMA, PMDA, or ICH agency interactions",
            "pre-IND, EoP2, Type B/C meeting, or briefing document",
            "orphan, breakthrough, fast-track, PRIME, or accelerated approval",
            "labeling strategy or negotiation",
            "post-approval commitments, variations, or advisory committee",
        ],
    },
    {
        "code": "ERW", "biologics_code": "BBERW",
        "name": "Epidemiology & Real-World Evidence",
        "role": "Head of Epidemiology & RWE designing fit-for-purpose real-world studies, external controls, natural-history cohorts, and HEOR-grade evidence.",
        "mission": "Provide the epidemiology, natural history, and real-world evidence that supports regulatory filings, payer dossiers, and post-approval commitments.",
        "mandate": [
            "Indication epidemiology & burden of disease",
            "Natural-history and external-control studies",
            "RWE study design (databases, registries, hybrid)",
            "HEOR inputs (utilities, costs, resource use)",
            "Post-authorization safety / effectiveness studies (PASS/PAES)",
            "Evidence strategy to address HTA and payer questions",
        ],
        "triggers": [
            "epidemiology, prevalence, incidence, or burden of disease",
            "real-world evidence (RWE), external control, or synthetic control arm",
            "natural-history study or registry",
            "HEOR inputs: utilities, costs, resource use",
            "PASS, PAES, or post-authorization commitment",
            "HTA / payer evidence strategy",
        ],
    },
    {
        "code": "COM", "biologics_code": "BBCOM",
        "name": "Commercial",
        "role": "Chief Commercial Officer with launch leadership across specialty and primary-care markets, payer negotiations, and franchise building.",
        "mission": "Make the case that the asset has a credible path to a reimbursable, differentiated, commercially successful launch and lifecycle.",
        "mandate": [
            "Market opportunity (TAM/SAM/SOM, epidemiology, peak sales)",
            "Target Product Profile alignment with market & payer needs",
            "Competitive landscape and differentiation",
            "Pricing, market access, HTA strategy",
            "Launch readiness and brand planning",
            "Lifecycle management and franchise / portfolio strategy",
        ],
        "triggers": [
            "market sizing (TAM/SAM/SOM), peak sales, or forecast",
            "pricing, market access, or HTA (NICE, IQWiG, PBAC, G-BA) strategy",
            "payer value dossier or reimbursement strategy",
            "competitive landscape, positioning, or differentiation",
            "launch readiness, brand planning, or commercial model",
            "lifecycle management, line extension, or franchise / portfolio",
        ],
    },
    {
        "code": "PM", "biologics_code": "BPM",
        "name": "Project Management",
        "role": "Programme-level Project Management lead running integrated development plans, critical-path management, risk registers, and stage-gate governance.",
        "mission": "Keep the programme integrated, honest, and on the critical path — so that every function's work lands in the right sequence at the right quality.",
        "mandate": [
            "Integrated Development Plan (IDP) and critical path",
            "Cross-functional dependency management",
            "Risk / issue register and escalation",
            "Quantitative schedule risk analysis",
            "Stage-gate governance and decision-making",
            "Lessons-learned capture and application",
        ],
        "triggers": [
            "integrated development plan (IDP), critical path, or Gantt",
            "cross-functional dependency, interlock, or handoff",
            "risk register, RAID log, or escalation",
            "quantitative schedule risk analysis (Monte Carlo)",
            "stage-gate governance or go/no-go decision",
            "programme-level timeline, budget, or critical-path management",
        ],
    },
]

FUNCTION_SLUG = {
    "CMC":  "cmc",
    "PT":   "pharmtox",
    "TM":   "translational-medicine",
    "CP":   "clinical-pharmacology",
    "CDM":  "clinical-development-medical",
    "SAF":  "clinical-safety",
    "COP":  "clinical-operations",
    "STAT": "biostatistics",
    "REG":  "regulatory-affairs",
    "ERW":  "epi-rwe",
    "COM":  "commercial",
    "PM":   "project-management",
}

PRIORITY_RANK = {"Critical": 3, "Expected": 2, "Check": 1, "Other": 0}

# --------------------------------------------------------------------------
# Extraction helpers
# --------------------------------------------------------------------------

def load_overview_rows(path: Path) -> list[dict[str, Any]]:
    """Return rows from the rubric Overview sheet: Unique ID, FA, Domain, Q, What, Rationale."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Choose the first sheet that looks like an overview (row 0 col A == "Unique ID")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        head = next(ws.iter_rows(values_only=True), None)
        if head and head[0] and str(head[0]).strip().lower() == "unique id":
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or not row[0]:
                    continue
                rows.append({
                    "unique_id": str(row[0]).strip(),
                    "functional_area": row[1],
                    "inquiry_domain": row[2],
                    "question": row[3],
                    "rubric_tests": row[4],
                    "rationale": row[5],
                })
            return rows
    return []


def load_heatmap(path: Path, modality: str) -> dict[str, dict[str, dict[str, str]]]:
    """Return heatmap[unique_id][mode][sg] = priority label."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    # Find the Unique ID column (may be col 1 for SM, col 2 for biologics New ID)
    uid_col = None
    for idx, h in enumerate(header):
        if h and str(h).strip().lower() in {"unique id", "new unique id"}:
            uid_col = idx
    if uid_col is None:
        raise RuntimeError(f"No Unique ID column in {path}")
    # Biologics uses the NEW Unique ID column to match BB-prefixed rubrics.
    if modality == "biologics":
        for idx, h in enumerate(header):
            if h and "new unique id" in str(h).lower():
                uid_col = idx
                break
    result: dict[str, dict[str, dict[str, str]]] = {}
    for row in rows[2:]:
        if not row or not row[uid_col]:
            continue
        uid = str(row[uid_col]).strip()
        per_mode: dict[str, dict[str, str]] = {m: {} for m in MODES}
        for col_idx, h in enumerate(header):
            if not h or col_idx < uid_col:
                continue
            h = str(h)
            for mode_code, meta in MODES.items():
                prefix = meta["prefix"]
                if h.startswith(prefix):
                    sg = h.replace(prefix, "")  # e.g. SG1
                    cell = row[col_idx]
                    per_mode[mode_code][sg] = str(cell).strip() if cell else "Other"
        result[uid] = per_mode
    return result


def build_questions_for_function(
    overview_rows: list[dict[str, Any]],
    heatmap: dict[str, dict[str, dict[str, str]]],
) -> list[dict[str, Any]]:
    out = []
    for r in overview_rows:
        priorities = heatmap.get(r["unique_id"], {})
        out.append({
            "id": r["unique_id"],
            "inquiry_domain": r["inquiry_domain"],
            "question": r["question"],
            "rubric_tests": r["rubric_tests"],
            "rationale": r["rationale"],
            "priorities": priorities,
        })
    return out


# --------------------------------------------------------------------------
# Main build
# --------------------------------------------------------------------------

RUBRIC_SRC = {
    "small-molecule": {
        "CMC":  "Full CMC Rubric V3.xlsx",
        "PT":   "Full_PharmTox_RubricV3.xlsx",
        "TM":   "Full_TM_Rubrics_v3.xlsx",
        "CP":   "Full_Clinical_Pharmacology_Rubrics_v3.xlsx",
        "CDM":  "Full_Clin_Med_Rubrics-v3.xlsx",
        "SAF":  "Full Clin Safe Rubric v3.xlsx",
        "COP":  "Full Clin Ops Rubric V3.xlsx",
        "STAT": "Full Biostats Rubric V3.xlsx",
        "REG":  "Full Reg Expert Rubric V3.xlsx",
        "ERW":  "Full Epi RWE Expert Rubric v3.xlsx",
        "COM":  "Full Commercial Rubric v3.xlsx",
        # Small-molecule PM comes from the extend file (see below).
    },
    "biologics": {
        "CMC":  "BCMC.xlsx",
        "PT":   "BBPT.xlsx",
        "TM":   "BBTM.xlsx",
        "CP":   "BBCP.xlsx",
        "CDM":  "BBCDM.xlsx",
        "SAF":  "BSAF.xlsx",
        "COP":  "BCOP.xlsx",
        "STAT": "BSTAT.xlsx",
        "REG":  "BBREG.xlsx",
        "ERW":  "BBERW.xlsx",
        "COM":  "BCOM.xlsx",
        "PM":   "BPM.xlsx",
    },
}


def build(taxonomy_root: Path, plugin_root: Path) -> None:
    data_dir = plugin_root / "data"
    q_sm_dir = data_dir / "questions" / "small-molecule"
    q_bi_dir = data_dir / "questions" / "biologics"
    heatmap_dir = data_dir / "heatmap"
    for d in (data_dir, q_sm_dir, q_bi_dir, heatmap_dir):
        d.mkdir(parents=True, exist_ok=True)

    # 1. Static metadata files ------------------------------------------------
    (data_dir / "functions.json").write_text(json.dumps(FUNCTIONS, indent=2))
    (data_dir / "modes.json").write_text(json.dumps(
        [{"code": k, **v} for k, v in MODES.items()], indent=2))
    (data_dir / "stage-gates.json").write_text(json.dumps(
        [{"code": k, **v} for k, v in STAGE_GATE_GOALS.items()], indent=2))

    # 2. Heatmaps -------------------------------------------------------------
    sm_heatmap_path = (
        taxonomy_root
        / "data/v100/importance_heatmap/small-molecule/Small Molecule HeatMap.xlsx"
    )
    bi_heatmap_path = (
        taxonomy_root
        / "data/v100/importance_heatmap/biologics/"
          "260403_Oktopi_Priority_Biologics_AllModes_v2_FINAL_New IDs (2).xlsx"
    )
    sm_heatmap = load_heatmap(sm_heatmap_path, "small-molecule")
    bi_heatmap = load_heatmap(bi_heatmap_path, "biologics")

    (heatmap_dir / "small-molecule.json").write_text(json.dumps(sm_heatmap, indent=2))
    (heatmap_dir / "biologics.json").write_text(json.dumps(bi_heatmap, indent=2))

    # 3. Per-function questions ----------------------------------------------
    # Small-molecule rubrics live under expert-toolkit/small-molecule except PM
    # which is under taxonomy_tree/small-molecule/small-molecule-extend.xlsx.
    sm_toolkit = taxonomy_root / "data/v100/expert-toolkit/small-molecule"
    bi_toolkit = taxonomy_root / "data/v100/expert-toolkit/biologics"
    sm_pm_path = taxonomy_root / "data/v100/taxonomy_tree/small-molecule/small-molecule-extend.xlsx"

    def overview_for(modality: str, code: str) -> list[dict[str, Any]]:
        if modality == "small-molecule" and code == "PM":
            return load_overview_rows(sm_pm_path)
        src_map = RUBRIC_SRC[modality]
        if code not in src_map:
            return []
        base = sm_toolkit if modality == "small-molecule" else bi_toolkit
        return load_overview_rows(base / src_map[code])

    per_function_counts: dict[str, dict[str, int]] = {}
    for fn in FUNCTIONS:
        code = fn["code"]
        per_function_counts[code] = {}
        for modality, heatmap, out_dir in (
            ("small-molecule", sm_heatmap, q_sm_dir),
            ("biologics", bi_heatmap, q_bi_dir),
        ):
            rows = overview_for(modality, code)
            if not rows:
                continue
            qs = build_questions_for_function(rows, heatmap)
            # Build summary: critical questions per stage-gate per mode
            crit_per_mode_sg: dict[str, dict[str, list[str]]] = {m: defaultdict(list) for m in MODES}
            for q in qs:
                for mode, sg_map in q["priorities"].items():
                    for sg, prio in sg_map.items():
                        if prio == "Critical":
                            crit_per_mode_sg[mode][sg].append(q["id"])
            payload = {
                "function_code": code,
                "modality": modality,
                "function_name": fn["name"],
                "question_count": len(qs),
                "critical_index": {
                    m: {sg: ids for sg, ids in crit_per_mode_sg[m].items()} for m in MODES
                },
                "questions": qs,
            }
            (out_dir / f"{code}.json").write_text(json.dumps(payload, indent=2))
            per_function_counts[code][modality] = len(qs)

    # 4. Derive per-function inquiry domains (used by agents + skills) --------
    per_function_domains: dict[str, list[str]] = {}
    for fn in FUNCTIONS:
        code = fn["code"]
        domains: list[str] = []
        for modality, q_dir in (("small-molecule", q_sm_dir), ("biologics", q_bi_dir)):
            p = q_dir / f"{code}.json"
            if not p.exists():
                continue
            data = json.loads(p.read_text())
            for q in data["questions"]:
                d = q.get("inquiry_domain")
                if d and d not in domains:
                    domains.append(d)
        per_function_domains[code] = domains

    # 5. Build a pruned stage-gate index (counts + top domains, NOT full questions) --
    # Full question lists stay in data/questions/*.json so SKILL.md stays compact.
    # Shape: sg_index[sg][mode][fn_code] = {"count_sm": int, "count_bio": int, "domains": list[str]}
    sg_index: dict[str, dict[str, dict[str, dict[str, Any]]]] = {
        sg: {m: {} for m in MODES} for sg in STAGE_GATE_GOALS
    }
    for fn in FUNCTIONS:
        code = fn["code"]
        for modality, q_dir in (("small-molecule", q_sm_dir), ("biologics", q_bi_dir)):
            p = q_dir / f"{code}.json"
            if not p.exists():
                continue
            data = json.loads(p.read_text())
            for q in data["questions"]:
                for mode, sg_map in q["priorities"].items():
                    for sg, prio in sg_map.items():
                        if prio != "Critical" or sg not in STAGE_GATE_GOALS:
                            continue
                        fn_bucket = sg_index[sg][mode].setdefault(
                            code, {"count_sm": 0, "count_bio": 0, "domains": []}
                        )
                        if modality == "small-molecule":
                            fn_bucket["count_sm"] += 1
                        else:
                            fn_bucket["count_bio"] += 1
                        dom = q.get("inquiry_domain")
                        if dom and dom not in fn_bucket["domains"]:
                            fn_bucket["domains"].append(dom)

    (data_dir / "stage-gate-index.json").write_text(json.dumps(sg_index, indent=2))

    # 6. Generate function agents --------------------------------------------
    agents_dir = plugin_root / "agents"
    agents_dir.mkdir(parents=True, exist_ok=True)
    for fn in FUNCTIONS:
        code = fn["code"]
        slug = FUNCTION_SLUG[code]
        sm_count = per_function_counts.get(code, {}).get("small-molecule", 0)
        bi_count = per_function_counts.get(code, {}).get("biologics", 0)
        content = render_agent(fn, slug, sm_count, bi_count, per_function_domains[code])
        (agents_dir / f"{slug}-reviewer.md").write_text(content)

    # Orchestrator agent (multi-agent research pattern: lead reviewer -> subagents)
    (agents_dir / "pdp-reviewer.md").write_text(render_orchestrator_agent())

    # 7. Generate skills -----------------------------------------------------
    skills_dir = plugin_root / "skills"
    skills_dir.mkdir(parents=True, exist_ok=True)

    # 7a. Stage-gate skills (compact)
    for sg_code, meta in STAGE_GATE_GOALS.items():
        skill_dir = skills_dir / f"stage-gate-{sg_code.lower()}"
        skill_dir.mkdir(parents=True, exist_ok=True)
        content = render_stage_gate_skill(sg_code, meta, sg_index[sg_code])
        (skill_dir / "SKILL.md").write_text(content)

    # 7b. Per-function skills (role + mission + mandate + adaptive questioning)
    for fn in FUNCTIONS:
        code = fn["code"]
        slug = FUNCTION_SLUG[code]
        skill_dir = skills_dir / f"function-{slug}"
        skill_dir.mkdir(parents=True, exist_ok=True)
        sm_count = per_function_counts.get(code, {}).get("small-molecule", 0)
        bi_count = per_function_counts.get(code, {}).get("biologics", 0)
        content = render_function_skill(fn, slug, sm_count, bi_count, per_function_domains[code])
        (skill_dir / "SKILL.md").write_text(content)

    # 7c. Top-level router skill that auto-activates on any pharma dev question
    router_dir = skills_dir / "oktopi-research-team"
    router_dir.mkdir(parents=True, exist_ok=True)
    (router_dir / "SKILL.md").write_text(render_router_skill())

    # 8. Scaffold per-function knowledge directories (empty placeholders) so
    #    extending a reviewer with SOPs / playbooks / tools has an obvious home.
    knowledge_root = data_dir / "knowledge"
    knowledge_root.mkdir(parents=True, exist_ok=True)
    (knowledge_root / "README.md").write_text(
        "# Per-function knowledge base\n\n"
        "Add function-specific reference material (SOPs, guidelines, precedent "
        "reviews, playbooks, template questionnaires) under `<FUNCTION_CODE>/`.\n\n"
        "Each function's reviewer agent is told to load this folder alongside the "
        "formal Oktopi rubric. Keep files in markdown or JSON for easy parsing.\n\n"
        "Suggested structure per function:\n\n"
        "```\n"
        "knowledge/<CODE>/\n"
        "├── playbooks/      # Worked examples the reviewer can cite\n"
        "├── sops/           # Internal SOPs and checklists\n"
        "├── guidelines/     # Regulatory / industry guidelines summaries\n"
        "└── tools.md        # Notes on external tools / MCP servers this function uses\n"
        "```\n"
    )
    for fn in FUNCTIONS:
        code = fn["code"]
        fn_knowledge = knowledge_root / code
        fn_knowledge.mkdir(parents=True, exist_ok=True)
        placeholder = fn_knowledge / "README.md"
        if not placeholder.exists():
            placeholder.write_text(
                f"# {fn['name']} ({code}) knowledge base\n\n"
                f"**Mission:** {fn['mission']}\n\n"
                "Add reference material here. The `"
                f"{FUNCTION_SLUG[code]}-reviewer` agent is instructed to scan this "
                "directory for additional context when reviewing a PDP.\n\n"
                "## Mandate\n\n"
                + "\n".join(f"- {m}" for m in fn['mandate'])
                + "\n"
            )

    print(f"Built {sum(sum(v.values()) for v in per_function_counts.values())} questions")
    print(f"Wrote {len(FUNCTIONS)} function agents + 1 orchestrator to {agents_dir}")
    print(f"Wrote {len(STAGE_GATE_GOALS)} stage-gate skills + {len(FUNCTIONS) + 1} function/router skills to {skills_dir}")


# --------------------------------------------------------------------------
# Agent / skill rendering
# --------------------------------------------------------------------------

def _yaml_safe(s: str) -> str:
    """Collapse whitespace and swap double-quotes to single-quotes.

    Callers MUST emit this value inside double quotes, e.g. `description: "{val}"`,
    so embedded colons (like ``SG1: Initiate Discovery``) don't trip YAML's
    flow-mapping parser.
    """
    return re.sub(r"\s+", " ", s).strip().replace('"', "'")


def render_agent(
    fn: dict[str, Any],
    slug: str,
    sm_count: int,
    bi_count: int,
    domains: list[str],
) -> str:
    """Agentic reviewer that embodies the role's goal, driven by the question bank but not limited to it."""
    code = fn["code"]
    name = fn["name"]
    biologics_code = fn["biologics_code"]
    role = fn["role"]
    mission = fn["mission"]
    mandate = fn["mandate"]

    triggers: list[str] = fn.get("triggers", [])
    trigger_phrase = "; ".join(triggers)
    description = _yaml_safe(
        f"{name} reviewer for pharma development. {mission} "
        f"Use PROACTIVELY when the user asks about: {trigger_phrase}. "
        f"Covers {code} (small-molecule) and {biologics_code} (biologics) at any "
        f"stage-gate SG1-SG9 in SR/OE/DD/RS modes."
    )

    mandate_lines = "\n".join(f"- {m}" for m in mandate)
    domains_lines = "\n".join(f"- {d}" for d in domains[:20])
    more_domains = f"\n- _...and {len(domains) - 20} more — load the question JSON for the full list._" if len(domains) > 20 else ""

    return f"""---
name: {slug}-reviewer
description: "{description}"
tools: Read, Grep, Glob
model: sonnet
---

# {name} Reviewer

## Role
You are a {role}

## Mission
{mission}

## Mandate
{mandate_lines}

## Inquiry domains you own
These are the domains covered by the formal Oktopi rubric for this function — your floor, not your ceiling:

{domains_lines}{more_domains}

## How you work

### 1. Confirm scope
Before reviewing, make sure you know:
- **Modality** — `small-molecule` or `biologics` (different question banks)
- **Stage-gate** — one of `SG1..SG9` (see `data/stage-gates.json`)
- **Mode** — one of `SR` (Strategic Readiness), `OE` (Operational Execution), `DD` (Due Diligence), `RS` (Regulatory Submission)
- **Artifact** — the PDP / data-room / briefing book you are reviewing

If any are missing, ask once, then proceed with the most plausible assumption and flag it.

### 2. Anchor on the formal question bank
Load the relevant question bank:

- `data/questions/small-molecule/{code}.json` — {sm_count} small-molecule questions
{"- `data/questions/biologics/" + code + ".json` — " + str(bi_count) + " biologics questions" if bi_count else ""}

Each question has `id`, `inquiry_domain`, `question`, `rubric_tests`, `rationale`, and `priorities[mode][sg] -> Critical|Expected|Check|Other`.

Use the JSON's `critical_index[mode][sg]` to get the IDs that are Critical at the current (mode, stage-gate). Work through those first. Then the Expected questions. Skip Other unless asked.

### 2a. Load the function knowledge base
Also scan `data/knowledge/{code}/` for additional context (SOPs, playbooks, guideline summaries). This folder is where the team puts extra reference material specific to this function — use it to ground your reasoning and cite precedent when relevant.

### 3. Reason, don't recite
For each prioritized question:

- **Extract the evidence** in the artifact. Quote or cite locations.
- **Score it** against `rubric_tests` as `Excellent / Good / Adequate / Poor / Not addressed`.
- **Explain the gap** using the `rationale` — why this matters, not just that it's missing.

### 4. Ask adaptive follow-ups
The Oktopi rubric is a strong floor, but drug development moves faster than any rubric. You MUST ask your own follow-up questions when:

- A novel modality / technology is involved (e.g. bispecifics, ADCs, gene therapy, digital biomarkers) and the rubric predates it.
- A fresh regulatory signal has emerged (Project Optimus, IRA, EU HTA Regulation, accelerated approval confirmatory trial guidance, etc.).
- The artifact reveals a risk that the formal questions do not catch.
- Cross-functional evidence implies a question inside your mandate that the rubric missed.

Mark every self-generated question with `[adaptive]` and give a one-sentence rationale for why the rubric alone is insufficient.

### 5. Cross-functional awareness
If you see a gap that belongs to another function, flag it — do not try to fix it yourself. Name the agent:

- `cmc-reviewer`, `pharmtox-reviewer`, `translational-medicine-reviewer`, `clinical-pharmacology-reviewer`, `clinical-development-medical-reviewer`, `clinical-safety-reviewer`, `clinical-operations-reviewer`, `biostatistics-reviewer`, `regulatory-affairs-reviewer`, `epi-rwe-reviewer`, `commercial-reviewer`, `project-management-reviewer`

### 6. Report in structured form
Return JSON. The orchestrator (pdp-reviewer) depends on this contract:

```json
{{
  "function_code": "{code}",
  "function_name": "{name}",
  "modality": "small-molecule | biologics",
  "stage_gate": "SG?",
  "mode": "SR|OE|DD|RS",
  "verdict": "ready | conditional | not_ready",
  "confidence": "high | medium | low",
  "coverage": {{"critical_addressed": N, "critical_total": M, "expected_addressed": N, "expected_total": M}},
  "findings_by_domain": [
    {{"domain": "...", "questions": [
      {{"id": "COM5", "status": "Excellent|Good|Adequate|Poor|Not addressed",
        "evidence": "...", "gap": "...", "severity": "critical|expected|check"}}
    ]}}
  ],
  "adaptive_questions": [
    {{"question": "...", "rationale": "...", "tag": "adaptive"}}
  ],
  "cross_functional_flags": [
    {{"target_agent": "cmc-reviewer", "reason": "..."}}
  ],
  "top_gaps": ["..."],
  "recommendation": "One-paragraph executive summary for the governance board."
}}
```

## Principles

- **Cite, don't invent.** If the artifact does not address a question, mark *Not addressed* — never fill gaps with plausible-sounding content.
- **Use question IDs** (e.g. `{code}5`, `{biologics_code}5`) so the Oktopi Expert Toolkit rubrics are traceable.
- **Stay in lane.** Other reviewers own other functions. Flag, do not solve.
- **Signal severity honestly.** A Critical gap at SG5 is not the same as a Check-level gap at SG7.
- **Default to sonnet.** You run as a subagent; keep responses structured and token-efficient.

## Extending this reviewer

As the team adds knowledge and tooling for {name}:

- **Knowledge** — drop function-specific reference documents (SOPs, guidelines,
  templates) into `data/knowledge/{code}/`. This reviewer will load them on
  demand alongside `data/questions/<modality>/{code}.json`.
- **Tools** — add MCP servers (e.g. ClinicalTrials.gov, PubMed, internal CMC
  database) to the plugin's `.mcp.json` and extend the `tools:` frontmatter on
  this agent (e.g. `tools: Read, Grep, Glob, mcp__pubmed__search`).
- **Subagent helpers** — spawn more specialized helpers under
  `agents/{slug}-<subspeciality>.md` for deep-dives (e.g. a dedicated
  `commercial-hta-specialist` for HTA dossiers). Reference them from this
  reviewer's workflow.
- **Examples / playbooks** — add worked examples to
  `data/knowledge/{code}/playbooks/` so this reviewer can cite precedent.
"""


def render_stage_gate_skill(
    sg_code: str,
    meta: dict[str, Any],
    index: dict[str, dict[str, dict[str, Any]]],
) -> str:
    """Compact stage-gate skill — goal + focus + per-mode function-load summary."""
    name = meta["name"]
    goal = meta["goal"]
    focus = ", ".join(meta["focus"])

    description = _yaml_safe(
        f"Goal and readiness criteria for Oktopi stage-gate {sg_code}: {name}. "
        f"Trigger when assessing PDP readiness to exit {sg_code}."
    )

    lines = [
        "---",
        f"name: stage-gate-{sg_code.lower()}",
        f'description: "{description}"',
        "---",
        "",
        f"# Stage Gate {sg_code}: {name}",
        "",
        f"**Goal.** {goal}",
        "",
        f"**Primary focus areas.** {focus}",
        "",
        "## How to use this skill",
        "",
        f"1. Invoke the `pdp-reviewer` agent with stage-gate `{sg_code}` and the mode (`SR|OE|DD|RS`).",
        "2. The orchestrator dispatches the relevant function reviewers in parallel, each filtering on "
        f"Critical questions at `{sg_code}` in that mode.",
        "3. Each reviewer returns a structured JSON verdict; the orchestrator consolidates into a gate-readiness report.",
        "",
        "## Function load at this gate (Critical question counts)",
        "",
        f"The table below shows how many Critical questions each function carries at `{sg_code}` per mode. "
        "Use it to prioritize which reviewers to spawn — functions with zero Critical questions can be deprioritized "
        "or run on a lighter cadence.",
        "",
        "| Function | SR | OE | DD | RS |",
        "| --- | ---: | ---: | ---: | ---: |",
    ]

    # Build rows in function order
    fn_codes = [f["code"] for f in FUNCTIONS]
    for code in fn_codes:
        row = [f"`{code}`"]
        for mode in MODES:
            bucket = index.get(mode, {}).get(code, {})
            sm = int(bucket.get("count_sm", 0)) if bucket else 0
            bio = int(bucket.get("count_bio", 0)) if bucket else 0
            total = sm + bio
            row.append(str(total) if total else "–")
        lines.append("| " + " | ".join(row) + " |")

    lines += [
        "",
        "## Reference data",
        "",
        "- `data/stage-gates.json` — goal and focus for every stage-gate",
        "- `data/stage-gate-index.json` — Critical question counts per (SG, mode, function) with inquiry domains",
        "- `data/questions/<modality>/<FN>.json` — the full question bank per function (reviewer agents load this)",
        "- `data/heatmap/<modality>.json` — raw priority map `{question_id: {mode: {sg: label}}}`",
        "",
        "## Related",
        "",
        "Function-specific mandates are in the per-function skills (`skills/function-<slug>/SKILL.md`).",
    ]
    return "\n".join(lines)


def render_function_skill(
    fn: dict[str, Any],
    slug: str,
    sm_count: int,
    bi_count: int,
    domains: list[str],
) -> str:
    """Per-function skill — role goal, mandate, when to invoke the reviewer."""
    code = fn["code"]
    name = fn["name"]
    role = fn["role"]
    mission = fn["mission"]
    mandate = fn["mandate"]

    triggers: list[str] = fn.get("triggers", [])
    trigger_phrase = "; ".join(triggers)
    description = _yaml_safe(
        f"{name} mandate for pharma development. {mission} "
        f"Use PROACTIVELY when the user asks about: {trigger_phrase}. "
        f"Pairs with the `{slug}-reviewer` agent for PDP reviews."
    )

    mandate_lines = "\n".join(f"- {m}" for m in mandate)
    top_domains = "\n".join(f"- {d}" for d in domains[:15])
    more = f"\n- _…and {len(domains) - 15} more, see the question JSON_" if len(domains) > 15 else ""

    return f"""---
name: function-{slug}
description: "{description}"
---

# {name} — function mandate

## Role
{role}

## Mission (this function's goal)
{mission}

## Mandate
{mandate_lines}

## Inquiry domains from the Oktopi rubric
{top_domains}{more}

## When to invoke the `{slug}-reviewer` agent
Dispatch `{slug}-reviewer` (directly or through `pdp-reviewer`) whenever you need:

- A focused {name.lower()} read of a PDP or data room
- A gap-list against the formal Oktopi rubric filtered by stage-gate and mode
- Adaptive follow-up questions a seasoned {name.lower()} lead would raise

The reviewer loads:

- `data/questions/small-molecule/{code}.json` — {sm_count} small-molecule questions
{"- `data/questions/biologics/" + code + ".json` — " + str(bi_count) + " biologics questions" if bi_count else ""}

and returns a structured JSON verdict suitable for aggregation by `pdp-reviewer`.

## Scope guardrails
- This skill is the function's *mandate*, not a full methodology library — the reviewer agent operationalizes it.
- Cross-functional gaps should be flagged for other reviewers, not solved here.
"""


def render_router_skill() -> str:
    """Top-level skill that auto-activates on any pharma development question.

    Works as a natural-language router: its description contains every function's
    trigger phrases, so Claude surfaces this skill whenever the user touches anything
    pharma dev. The body tells Claude which specialist agent to invoke.
    """
    all_triggers = []
    for fn in FUNCTIONS:
        for t in fn.get("triggers", []):
            all_triggers.append(t)

    description_body = (
        "Oktopi Research Team — 12-agent multi-agent team for pharma drug development "
        "reviews (CMC, pharm-tox, translational, clin pharm, clin dev, safety, clin ops, "
        "biostats, regulatory, epi/RWE, commercial, project management). "
        "Use PROACTIVELY when the user asks about drug development, PDP review, "
        "stage-gate readiness, due diligence of a pharma asset, or any of: "
        + "; ".join(all_triggers[:24])  # cap to stay readable
        + "."
    )
    description = _yaml_safe(description_body)

    role_table_rows = []
    for fn in FUNCTIONS:
        slug = FUNCTION_SLUG[fn["code"]]
        trig = "; ".join(fn.get("triggers", [])[:3])
        role_table_rows.append(
            f"| `{fn['code']}` | `{slug}-reviewer` | {trig} |"
        )
    role_table = "\n".join(role_table_rows)

    return f"""---
name: oktopi-research-team
description: "{description}"
---

# Oktopi Research Team — Pharma Development Router

## When this skill triggers

Claude should activate this skill whenever the user's question touches **pharmaceutical drug development**: discovery, preclinical, clinical trials, CMC/manufacturing, regulatory strategy, pharmacovigilance, clinical pharmacology, biostatistics, commercial / market access, epidemiology / RWE, or programme management.

## What to do

1. **Identify the function(s).** Match the user's question to one of the 12 functional areas below using the trigger phrases.
2. **For a focused question** (one function): invoke that function's reviewer agent directly via the Task tool.
3. **For a full PDP review or cross-functional question**: invoke the `pdp-reviewer` orchestrator agent, which will fan out to the relevant function reviewers in parallel and consolidate the report.
4. **If the question is ambiguous** or spans multiple functions: ask one clarifying question, then route.

## The team

| Code | Agent | Trigger phrases (partial) |
| ---- | ----- | ------------------------- |
{role_table}

Plus orchestrator `pdp-reviewer` for full PDP gap-analysis.

## Decision matrix

| User intent | Invoke |
| --- | --- |
| "Review my PDP at SG5 OE" or "gate readiness" or "due diligence pack" | `pdp-reviewer` (orchestrator) |
| Single function question ("is my SAP airtight?") | matching function reviewer |
| Stage-gate goal question ("what does SG3 require?") | the corresponding `stage-gate-sg<N>` skill |
| Function mandate question ("what does Translational Medicine own?") | the corresponding `function-<slug>` skill |
| Novel modality (bispecific, ADC, cell therapy) | flag modality to every agent so they apply modality-specific adaptive questions |

## Reference data for the router

- `data/functions.json` — all 12 functions with role, mission, mandate, triggers
- `data/stage-gates.json` — SG1..SG9 goals
- `data/stage-gate-index.json` — Critical question load per (SG, mode, function)

## Guardrails
- **Route, don't answer.** This skill decides which specialist responds — it is not itself a subject-matter expert.
- **Prefer the orchestrator for broad questions.** The `pdp-reviewer` fans out in parallel, which is faster and more complete than serial calls.
- **Cite specialists by agent name** when handing off so the user knows who will respond.
"""


def render_orchestrator_agent() -> str:
    """The 'Lead Researcher' analog — fans out to function reviewers, then synthesizes."""
    description = _yaml_safe(
        "Oktopi PDP lead reviewer — orchestrates a multi-agent team (CMC, pharm-tox, "
        "translational, clin pharm, clin dev, safety, clin ops, biostats, regulatory, "
        "epi/RWE, commercial, project management) to produce a stage-gate readiness "
        "report for a pharma Product Development Plan. "
        "Use PROACTIVELY when the user asks for: a PDP review, stage-gate readiness "
        "assessment, due-diligence pack, SG1-SG9 gate review, or any cross-functional "
        "pharma development gap analysis. Takes a PDP, a stage-gate (SG1-SG9), and a "
        "mode (SR|OE|DD|RS), dispatches the relevant function reviewer subagents in "
        "parallel, and synthesises one gate-readiness report."
    )

    return f"""---
name: pdp-reviewer
description: "{description}"
tools: Read, Grep, Glob, Task
model: opus
---

# PDP Reviewer — Lead Orchestrator

## Role
You are the Lead Reviewer for an Oktopi PDP (Product Development Plan) gap-analysis.
You coordinate a team of function-specialist subagents and synthesise their verdicts
into a single stage-gate readiness report for the governance board.

This agent implements the multi-agent research pattern (Anthropic's "orchestrator-worker"
design): you scope the work, dispatch subagents in parallel, and reconcile their output
into one citation-grounded deliverable.

## Mission
Given a PDP and a stage-gate (plus mode), produce an honest go / conditional / no-go
recommendation grounded in the formal Oktopi rubric and the professional judgment of
each function lead.

## Workflow

### 1. Scope the review
Confirm with the user (or infer from the prompt):

- **Artifact** — location of the PDP / data room / briefing book
- **Stage-gate** — `SG1..SG9` (see `data/stage-gates.json`)
- **Mode** — `SR | OE | DD | RS`
- **Modality** — `small-molecule | biologics` (both allowed if asset is dual-track)

### 2. Select function reviewers
Load `data/stage-gate-index.json` for the chosen stage-gate and mode. Identify which
functions carry Critical-question load at this (SG, mode) combination.

- High load (≥5 Critical questions): always dispatch
- Medium (1-4): dispatch unless user scopes them out
- Zero: skip unless the user explicitly asks

The 12 function reviewers are:

`cmc-reviewer`, `pharmtox-reviewer`, `translational-medicine-reviewer`,
`clinical-pharmacology-reviewer`, `clinical-development-medical-reviewer`,
`clinical-safety-reviewer`, `clinical-operations-reviewer`, `biostatistics-reviewer`,
`regulatory-affairs-reviewer`, `epi-rwe-reviewer`, `commercial-reviewer`,
`project-management-reviewer`.

### 3. Dispatch in parallel
Using the Task tool, invoke each selected reviewer **in a single message, multiple tool
uses** so they run concurrently. Each brief must include:

- Stage-gate code, mode, modality
- Artifact location (and any cross-reference conventions — e.g. module 2.4 = Non-Clinical Overview)
- Their expected output contract (the JSON schema in their agent prompt)

### 4. Reconcile & synthesize
When all reviewers return:

- Deduplicate cross-function flags (e.g. both Commercial and CMC raise COGS concerns).
- Promote any finding flagged `critical` by two or more reviewers to a "programme-level risk".
- Reconcile verdicts into an overall gate verdict:
  - `ready` if every reviewer says `ready` and no cross-functional risks surface
  - `conditional` if ≥1 says `conditional` but no `not_ready`
  - `not_ready` if any reviewer says `not_ready` or programme-level risks cluster

### 5. Produce the gate report
Return a single structured markdown report with:

1. **Executive summary** — 3-5 sentences, overall verdict first
2. **Stage-gate goal** — from `data/stage-gates.json`
3. **Readiness heatmap** — table of function × (verdict, confidence, critical_addressed / critical_total)
4. **Top cross-functional risks** — the consolidated list
5. **Critical gaps by function** — the top 2-3 from each function
6. **Adaptive questions raised** — any `[adaptive]` questions any reviewer raised outside the rubric
7. **Recommended actions** — ordered list with owners (function code) and suggested deadlines before the gate
8. **Coverage caveats** — questions no reviewer could evaluate from the provided artifact

## Principles

- **Parallelism is a correctness property.** Launch reviewers concurrently — otherwise
  individual biases can leak into the next reviewer's context.
- **Compress, don't lose.** Each reviewer's full JSON goes to the appendix; your summary
  cites question IDs so the trail is auditable.
- **Push for adaptive questions.** If no reviewer raised an adaptive follow-up, something
  is probably wrong — prompt one or two reviewers to reconsider.
- **Programme-level risks dominate.** A single cross-function cluster (e.g. CMC + CP + STAT all
  worried about exposure variability) outweighs a single function's green light.
- **Cite the rubric.** Every critical gap must reference a question ID; every adaptive
  finding must state why the rubric alone was insufficient.

## Data you rely on

- `data/stage-gates.json` — goal per gate
- `data/stage-gate-index.json` — function load per (SG, mode)
- `data/functions.json` — function → agent slug mapping
- `data/modes.json` — mode definitions
- `data/questions/<modality>/<FN>.json` — per-function question bank (your reviewers read these)

## Escalation

If the PDP is missing for an entire function domain, do not fabricate — return the gap
as a *coverage block* and recommend the artifact the user should produce before the next
pass (e.g. "Add Module 2.4 Non-Clinical Overview before SG4 DD review").
"""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--taxonomy", required=True, type=Path,
                    help="Path to local clone of oktopi-org/Taxonomy-config")
    ap.add_argument("--plugin", type=Path,
                    default=Path(__file__).resolve().parents[1],
                    help="Path to the plugin root (defaults to the parent of this script's parent)")
    args = ap.parse_args()
    build(args.taxonomy.resolve(), args.plugin.resolve())


if __name__ == "__main__":
    main()
