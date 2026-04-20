#!/usr/bin/env python3
"""Extract Oktopi taxonomy configuration into plugin data and agent/skill artifacts.

Reads from a local clone of oktopi-org/Taxonomy-config and writes:

    plugins/oktopi-dev/data/
        functions.json
        stage-gates.json
        modes.json
        heatmap/<modality>.json
        questions/<modality>/<FUNCTION_CODE>.json

    plugins/oktopi-dev/agents/<function>-reviewer.md      (one per function)
    plugins/oktopi-dev/skills/stage-gate-sg<N>/SKILL.md   (one per stage gate)

Run from the repo root:
    python3 plugins/oktopi-dev/scripts/build_taxonomy_data.py \
        --taxonomy ../Taxonomy-config

All data is regenerated from scratch each run; the produced files are the
runtime reference material used by the agents and skills.
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

# --------------------------------------------------------------------------
# Static metadata
# --------------------------------------------------------------------------
MODES = {
    "SR": {"name": "Strategic Readiness", "prefix": "C_"},
    "OE": {"name": "Operational Execution", "prefix": "OE_"},
    "DD": {"name": "Due Diligence", "prefix": "DD_"},
    "RS": {"name": "Regulatory Submission", "prefix": "RS_"},
}

STAGE_GATE_GOALS = {
    "SG1": {
        "name": "Initiate Discovery Program & Target Identification",
        "goal": "Validate a biologically credible target, a defensible rationale, and a "
                "screening plan that can yield testable hypotheses.",
        "focus": ["target validation", "MoA rationale", "IP freedom-to-operate",
                   "discovery plan & budget"],
    },
    "SG2": {
        "name": "Lead Nomination / Early Optimization",
        "goal": "Nominate a lead series with acceptable developability signals and a "
                "clear optimization plan to reach clinical candidate criteria.",
        "focus": ["lead criteria met", "DMPK/tox early signals", "scaffold IP",
                   "formulation feasibility"],
    },
    "SG3": {
        "name": "Clinical Candidate Selection / IND-Enabling Entry",
        "goal": "Confirm a single clinical candidate with an approved IND-enabling plan "
                "across CMC, tox, PK/PD, and clinical design.",
        "focus": ["candidate selection data", "IND-enabling study plan",
                   "CMC readiness for GLP/GMP", "target product profile draft"],
    },
    "SG4": {
        "name": "Pre-Clinical to FIH-Enabling (CMC, PK/PD, Tox Ready)",
        "goal": "Complete IND-enabling package: GLP tox complete, CMC GMP drug product "
                "and IND Module 3, FIH protocol and starting dose justification.",
        "focus": ["GLP tox package", "CMC GMP release", "FIH protocol", "starting dose"],
    },
    "SG5": {
        "name": "First-in-Human (FIH) Clinical Program Initiation",
        "goal": "Open IND, activate sites, dose the first subject safely and collect "
                "data suitable for dose-escalation and early PK/PD decisions.",
        "focus": ["IND clearance", "site activation", "FPFV safety",
                   "dose-escalation readiness"],
    },
    "SG6": {
        "name": "Proof of Concept (POC) & Phase 2 Readiness",
        "goal": "Demonstrate clinical proof of concept and confirm Phase 2 design, dose, "
                "population, endpoints, and operational feasibility.",
        "focus": ["POC criteria met", "Phase 2 protocol", "biomarker strategy",
                   "scale-up CMC"],
    },
    "SG7": {
        "name": "Confirmatory & Phase 3 Readiness (Reg & Commercial aligned)",
        "goal": "Lock Phase 3 design with regulatory and commercial alignment; confirm "
                "CMC registration strategy and site/CRO readiness.",
        "focus": ["Phase 3 design", "EoP2 alignment", "registration CMC", "TPP lock"],
    },
    "SG8": {
        "name": "Regulatory Submission / Payer Engagement / Pre-Launch",
        "goal": "Submit a reviewable NDA/BLA, finalize payer value dossier, and prepare "
                "commercial launch readiness.",
        "focus": ["NDA/BLA submission", "labeling strategy", "payer value dossier",
                   "launch readiness"],
    },
    "SG9": {
        "name": "Launch & Post-Market Lifecycle Management",
        "goal": "Execute a compliant launch and sustain post-market obligations, "
                "lifecycle planning, and RWE/safety monitoring.",
        "focus": ["PSUR/PBRER", "REMS/RMP", "LCM indications", "RWE commitments"],
    },
}

# Map function code (small-molecule) to canonical function name and biologics alias.
FUNCTIONS = [
    {"code": "CMC",  "biologics_code": "BBCMC",  "name": "Chemistry, Manufacturing, and Controls"},
    {"code": "PT",   "biologics_code": "BBPT",   "name": "Pharmacology & Toxicology"},
    {"code": "TM",   "biologics_code": "BBTM",   "name": "Translational Medicine"},
    {"code": "CP",   "biologics_code": "BBCP",   "name": "Clinical Pharmacology"},
    {"code": "CDM",  "biologics_code": "BBCDM",  "name": "Clinical Development / Medical"},
    {"code": "SAF",  "biologics_code": "BSAF",   "name": "Clinical Safety"},
    {"code": "COP",  "biologics_code": "BCOP",   "name": "Clinical Operations"},
    {"code": "STAT", "biologics_code": "BSTAT",  "name": "Biostatistics"},
    {"code": "REG",  "biologics_code": "BBREG",  "name": "Regulatory Affairs"},
    {"code": "ERW",  "biologics_code": "BBERW",  "name": "Epidemiology & Real-World Evidence"},
    {"code": "COM",  "biologics_code": "BBCOM",  "name": "Commercial"},
    {"code": "PM",   "biologics_code": "BPM",    "name": "Project Management"},
]

FUNCTION_SLUG = {
    "CMC":  "cmc",
    "PT":   "pharmtox",
    "TM":   "translational-medicine",
    "CP":   "clinical-pharmacology",
    "CDM":  "clinical-development-medical",
    "SAF":  "clinical-safety",
    "COP":  "clinical-operations",
    "STAT": "biostatistics",
    "REG":  "regulatory-affairs",
    "ERW":  "epi-rwe",
    "COM":  "commercial",
    "PM":   "project-management",
}

PRIORITY_RANK = {"Critical": 3, "Expected": 2, "Check": 1, "Other": 0}

# --------------------------------------------------------------------------
# Extraction helpers
# --------------------------------------------------------------------------

def load_overview_rows(path: Path) -> list[dict[str, Any]]:
    """Return rows from the rubric Overview sheet: Unique ID, FA, Domain, Q, What, Rationale."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Choose the first sheet that looks like an overview (row 0 col A == "Unique ID")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        head = next(ws.iter_rows(values_only=True), None)
        if head and head[0] and str(head[0]).strip().lower() == "unique id":
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or not row[0]:
                    continue
                rows.append({
                    "unique_id": str(row[0]).strip(),
                    "functional_area": row[1],
                    "inquiry_domain": row[2],
                    "question": row[3],
                    "rubric_tests": row[4],
                    "rationale": row[5],
                })
            return rows
    return []


def load_heatmap(path: Path, modality: str) -> dict[str, dict[str, dict[str, str]]]:
    """Return heatmap[unique_id][mode][sg] = priority label."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    # Find the Unique ID column (may be col 1 for SM, col 2 for biologics New ID)
    uid_col = None
    for idx, h in enumerate(header):
        if h and str(h).strip().lower() in {"unique id", "new unique id"}:
            uid_col = idx
    if uid_col is None:
        raise RuntimeError(f"No Unique ID column in {path}")
    # Biologics uses the NEW Unique ID column to match BB-prefixed rubrics.
    if modality == "biologics":
        for idx, h in enumerate(header):
            if h and "new unique id" in str(h).lower():
                uid_col = idx
                break
    result: dict[str, dict[str, dict[str, str]]] = {}
    for row in rows[2:]:
        if not row or not row[uid_col]:
            continue
        uid = str(row[uid_col]).strip()
        per_mode: dict[str, dict[str, str]] = {m: {} for m in MODES}
        for col_idx, h in enumerate(header):
            if not h or col_idx < uid_col:
                continue
            h = str(h)
            for mode_code, meta in MODES.items():
                prefix = meta["prefix"]
                if h.startswith(prefix):
                    sg = h.replace(prefix, "")  # e.g. SG1
                    cell = row[col_idx]
                    per_mode[mode_code][sg] = str(cell).strip() if cell else "Other"
        result[uid] = per_mode
    return result


def build_questions_for_function(
    overview_rows: list[dict[str, Any]],
    heatmap: dict[str, dict[str, dict[str, str]]],
) -> list[dict[str, Any]]:
    out = []
    for r in overview_rows:
        priorities = heatmap.get(r["unique_id"], {})
        out.append({
            "id": r["unique_id"],
            "inquiry_domain": r["inquiry_domain"],
            "question": r["question"],
            "rubric_tests": r["rubric_tests"],
            "rationale": r["rationale"],
            "priorities": priorities,
        })
    return out


# --------------------------------------------------------------------------
# Main build
# --------------------------------------------------------------------------

RUBRIC_SRC = {
    "small-molecule": {
        "CMC":  "Full CMC Rubric V3.xlsx",
        "PT":   "Full_PharmTox_RubricV3.xlsx",
        "TM":   "Full_TM_Rubrics_v3.xlsx",
        "CP":   "Full_Clinical_Pharmacology_Rubrics_v3.xlsx",
        "CDM":  "Full_Clin_Med_Rubrics-v3.xlsx",
        "SAF":  "Full Clin Safe Rubric v3.xlsx",
        "COP":  "Full Clin Ops Rubric V3.xlsx",
        "STAT": "Full Biostats Rubric V3.xlsx",
        "REG":  "Full Reg Expert Rubric V3.xlsx",
        "ERW":  "Full Epi RWE Expert Rubric v3.xlsx",
        "COM":  "Full Commercial Rubric v3.xlsx",
        # Small-molecule PM comes from the extend file (see below).
    },
    "biologics": {
        "CMC":  "BCMC.xlsx",
        "PT":   "BBPT.xlsx",
        "TM":   "BBTM.xlsx",
        "CP":   "BBCP.xlsx",
        "CDM":  "BBCDM.xlsx",
        "SAF":  "BSAF.xlsx",
        "COP":  "BCOP.xlsx",
        "STAT": "BSTAT.xlsx",
        "REG":  "BBREG.xlsx",
        "ERW":  "BBERW.xlsx",
        "COM":  "BCOM.xlsx",
        "PM":   "BPM.xlsx",
    },
}


def build(taxonomy_root: Path, plugin_root: Path) -> None:
    data_dir = plugin_root / "data"
    q_sm_dir = data_dir / "questions" / "small-molecule"
    q_bi_dir = data_dir / "questions" / "biologics"
    heatmap_dir = data_dir / "heatmap"
    for d in (data_dir, q_sm_dir, q_bi_dir, heatmap_dir):
        d.mkdir(parents=True, exist_ok=True)

    # 1. Static metadata files ------------------------------------------------
    (data_dir / "functions.json").write_text(json.dumps(FUNCTIONS, indent=2))
    (data_dir / "modes.json").write_text(json.dumps(
        [{"code": k, **v} for k, v in MODES.items()], indent=2))
    (data_dir / "stage-gates.json").write_text(json.dumps(
        [{"code": k, **v} for k, v in STAGE_GATE_GOALS.items()], indent=2))

    # 2. Heatmaps -------------------------------------------------------------
    sm_heatmap_path = (
        taxonomy_root
        / "data/v100/importance_heatmap/small-molecule/Small Molecule HeatMap.xlsx"
    )
    bi_heatmap_path = (
        taxonomy_root
        / "data/v100/importance_heatmap/biologics/"
          "260403_Oktopi_Priority_Biologics_AllModes_v2_FINAL_New IDs (2).xlsx"
    )
    sm_heatmap = load_heatmap(sm_heatmap_path, "small-molecule")
    bi_heatmap = load_heatmap(bi_heatmap_path, "biologics")

    (heatmap_dir / "small-molecule.json").write_text(json.dumps(sm_heatmap, indent=2))
    (heatmap_dir / "biologics.json").write_text(json.dumps(bi_heatmap, indent=2))

    # 3. Per-function questions ----------------------------------------------
    # Small-molecule rubrics live under expert-toolkit/small-molecule except PM
    # which is under taxonomy_tree/small-molecule/small-molecule-extend.xlsx.
    sm_toolkit = taxonomy_root / "data/v100/expert-toolkit/small-molecule"
    bi_toolkit = taxonomy_root / "data/v100/expert-toolkit/biologics"
    sm_pm_path = taxonomy_root / "data/v100/taxonomy_tree/small-molecule/small-molecule-extend.xlsx"

    def overview_for(modality: str, code: str) -> list[dict[str, Any]]:
        if modality == "small-molecule" and code == "PM":
            return load_overview_rows(sm_pm_path)
        src_map = RUBRIC_SRC[modality]
        if code not in src_map:
            return []
        base = sm_toolkit if modality == "small-molecule" else bi_toolkit
        return load_overview_rows(base / src_map[code])

    per_function_counts: dict[str, dict[str, int]] = {}
    for fn in FUNCTIONS:
        code = fn["code"]
        per_function_counts[code] = {}
        for modality, heatmap, out_dir in (
            ("small-molecule", sm_heatmap, q_sm_dir),
            ("biologics", bi_heatmap, q_bi_dir),
        ):
            rows = overview_for(modality, code)
            if not rows:
                continue
            qs = build_questions_for_function(rows, heatmap)
            # Build summary: critical questions per stage-gate per mode
            crit_per_mode_sg: dict[str, dict[str, list[str]]] = {m: defaultdict(list) for m in MODES}
            for q in qs:
                for mode, sg_map in q["priorities"].items():
                    for sg, prio in sg_map.items():
                        if prio == "Critical":
                            crit_per_mode_sg[mode][sg].append(q["id"])
            payload = {
                "function_code": code,
                "modality": modality,
                "function_name": fn["name"],
                "question_count": len(qs),
                "critical_index": {
                    m: {sg: ids for sg, ids in crit_per_mode_sg[m].items()} for m in MODES
                },
                "questions": qs,
            }
            (out_dir / f"{code}.json").write_text(json.dumps(payload, indent=2))
            per_function_counts[code][modality] = len(qs)

    # 4. Generate agent markdown ---------------------------------------------
    agents_dir = plugin_root / "agents"
    agents_dir.mkdir(parents=True, exist_ok=True)
    for fn in FUNCTIONS:
        code = fn["code"]
        slug = FUNCTION_SLUG[code]
        sm_path_rel = f"data/questions/small-molecule/{code}.json"
        bi_path_rel = f"data/questions/biologics/{code}.json"
        sm_count = per_function_counts.get(code, {}).get("small-molecule", 0)
        bi_count = per_function_counts.get(code, {}).get("biologics", 0)
        content = render_agent(fn, slug, sm_path_rel, bi_path_rel, sm_count, bi_count)
        (agents_dir / f"{slug}-reviewer.md").write_text(content)

    # 5. Generate stage-gate skills ------------------------------------------
    skills_dir = plugin_root / "skills"
    skills_dir.mkdir(parents=True, exist_ok=True)
    # Pre-compute: per stage-gate, per mode, list of (function_code, question_id, question_text)
    sg_index: dict[str, dict[str, list[tuple[str, str, str, str]]]] = {
        sg: {m: [] for m in MODES} for sg in STAGE_GATE_GOALS
    }

    def index_modality(modality: str, q_dir: Path) -> None:
        for fn in FUNCTIONS:
            code = fn["code"]
            p = q_dir / f"{code}.json"
            if not p.exists():
                continue
            data = json.loads(p.read_text())
            for q in data["questions"]:
                for mode, sg_map in q["priorities"].items():
                    for sg, prio in sg_map.items():
                        if prio == "Critical":
                            sg_index[sg][mode].append(
                                (modality, code, q["id"], q["question"])
                            )

    index_modality("small-molecule", q_sm_dir)
    index_modality("biologics", q_bi_dir)

    for sg_code, meta in STAGE_GATE_GOALS.items():
        skill_dir = skills_dir / f"stage-gate-{sg_code.lower()}"
        skill_dir.mkdir(parents=True, exist_ok=True)
        content = render_stage_gate_skill(sg_code, meta, sg_index[sg_code])
        (skill_dir / "SKILL.md").write_text(content)

    print(f"Built {sum(sum(v.values()) for v in per_function_counts.values())} questions")
    print(f"Wrote {len(FUNCTIONS)} agents to {agents_dir}")
    print(f"Wrote {len(STAGE_GATE_GOALS)} stage-gate skills to {skills_dir}")


# --------------------------------------------------------------------------
# Agent / skill rendering
# --------------------------------------------------------------------------

def render_agent(
    fn: dict[str, Any],
    slug: str,
    sm_path: str,
    bi_path: str,
    sm_count: int,
    bi_count: int,
) -> str:
    code = fn["code"]
    name = fn["name"]
    biologics_code = fn["biologics_code"]
    desc_parts = [
        f"{name} reviewer for Oktopi PDP gap-analysis.",
        f"Evaluates gap-analysis questions for function {code} (small-molecule)",
    ]
    if bi_count:
        desc_parts.append(f"and {biologics_code} (biologics)")
    desc_parts.append(
        "against the 9 stage-gates across Strategic Readiness, Operational Execution, "
        "Due Diligence, and Regulatory Submission modes."
    )
    description = " ".join(desc_parts)
    # YAML-safe one-line description
    description = re.sub(r"\s+", " ", description).strip()
    lines = [
        "---",
        f"name: {slug}-reviewer",
        f"description: {description}",
        "tools: Read, Grep, Glob",
        "---",
        "",
        f"# {name} Reviewer",
        "",
        f"You are a senior {name} expert reviewing a Product Development Plan (PDP) "
        f"for Oktopi. Your role covers function code **{code}** "
        f"(small-molecule) and **{biologics_code}** (biologics).",
        "",
        "## Your knowledge base",
        "",
        "Load these JSON files before responding (use the Read tool with the path "
        "relative to the plugin root):",
        "",
        f"- `{sm_path}` — {sm_count} small-molecule gap-analysis questions",
    ]
    if bi_count:
        lines.append(f"- `{bi_path}` — {bi_count} biologics gap-analysis questions")
    lines += [
        "- `data/stage-gates.json` — the 9 stage-gate goals (SG1–SG9)",
        "- `data/modes.json` — the 4 assessment modes (SR, OE, DD, RS)",
        "",
        "Each question entry has: `id`, `inquiry_domain`, `question`, `rubric_tests`, "
        "`rationale`, and a `priorities` map of `{mode -> {SGn -> Critical|Expected|Check|Other}}`.",
        "",
        "## How to review",
        "",
        "1. **Scope**: Ask (or infer from context) which modality (small-molecule vs. "
        "biologics), which stage-gate (SG1–SG9), and which mode (SR, OE, DD, RS) the "
        "user is reviewing against.",
        "2. **Prioritize**: Start with questions rated `Critical` for that (mode, SG) "
        "pair, then `Expected`, then `Check`. Skip `Other` unless asked.",
        "3. **Evaluate**: For each prioritized question, extract the evidence from the "
        "user-supplied document (PDP, slide deck, briefing book, etc.) and score it "
        "against the `rubric_tests` criteria. Call out gaps using the `rationale`.",
        "4. **Report**: Produce a structured summary per Inquiry Domain. For each "
        "question list:",
        "   - Question ID and text",
        "   - Evidence found (with source citation if available)",
        "   - Gap / risk if evidence is missing or weak",
        "   - Red-flag severity (Critical / Expected / Check)",
        "",
        "## Reporting contract",
        "",
        "End every review with:",
        "",
        "- A **go / no-go recommendation** for this function at the stage-gate",
        "- The top 3 **critical gaps** with owner suggestions",
        "- Questions you **could not evaluate** from available evidence",
        "",
        "## Guardrails",
        "",
        "- Never fabricate evidence. If the document does not address a question, mark "
        "it as *Not addressed* and surface it as a gap.",
        "- Stay within your functional area. If you see a gap in another function, flag "
        "it and recommend the relevant reviewer agent (see `data/functions.json`).",
        "- Cite question IDs (e.g., `COM5`, `BBCOM5`) so downstream tooling can link to "
        "the full rubric in the Oktopi Expert Toolkit.",
        "",
    ]
    return "\n".join(lines)


def render_stage_gate_skill(
    sg_code: str,
    meta: dict[str, Any],
    per_mode_questions: dict[str, list[tuple[str, str, str, str]]],
) -> str:
    name = meta["name"]
    goal = meta["goal"]
    focus = ", ".join(meta["focus"])

    description = (
        f"Stage-gate {sg_code} review goal for an Oktopi PDP: {name}. "
        f"Use when assessing readiness to exit {sg_code} across CMC, clinical, "
        f"regulatory, and commercial functions."
    )

    def bullets_for(mode: str) -> list[str]:
        items = per_mode_questions.get(mode, [])
        if not items:
            return ["- _No questions flagged Critical for this mode at this stage-gate._"]
        # Group by function code
        by_fn: dict[str, list[tuple[str, str, str, str]]] = defaultdict(list)
        for item in items:
            _, fn_code, qid, _ = item
            by_fn[fn_code].append(item)
        out = []
        for fn_code in sorted(by_fn):
            out.append(f"- **{fn_code}**")
            for modality, _, qid, qtext in sorted(by_fn[fn_code], key=lambda x: x[2]):
                qtext_short = re.sub(r"\s+", " ", str(qtext or "")).strip()
                if len(qtext_short) > 180:
                    qtext_short = qtext_short[:177] + "..."
                tag = "SM" if modality == "small-molecule" else "BIO"
                out.append(f"  - `{qid}` *({tag})* — {qtext_short}")
        return out

    lines = [
        "---",
        f"name: Stage Gate {sg_code} — {name}",
        f"description: {description}",
        "---",
        "",
        f"# Stage Gate {sg_code}: {name}",
        "",
        f"**Goal.** {goal}",
        "",
        f"**Primary focus areas.** {focus}",
        "",
        "## How to use",
        "",
        f"1. Confirm the modality (small-molecule or biologics) and mode (SR, OE, DD, RS) in scope.",
        f"2. For each function, run the corresponding function reviewer agent (`<function>-reviewer`).",
        f"3. Feed each agent this stage-gate code ({sg_code}) and mode so it filters on "
        f"Critical questions first.",
        f"4. Consolidate gaps per function into a single go/no-go decision for {sg_code}.",
        "",
        "## Critical questions at this stage-gate, by mode",
        "",
    ]
    for mode_code, mode_meta in MODES.items():
        lines.append(f"### {mode_code} — {mode_meta['name']}")
        lines.append("")
        lines.extend(bullets_for(mode_code))
        lines.append("")

    lines += [
        "## Data sources",
        "",
        "- `data/stage-gates.json` — goal and focus for every stage-gate",
        "- `data/questions/small-molecule/<FN>.json` — full rubric per small-molecule question",
        "- `data/questions/biologics/<FN>.json` — full rubric per biologics question",
        "- `data/heatmap/<modality>.json` — priority map `{question_id: {mode: {sg: label}}}`",
        "",
    ]
    return "\n".join(lines)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--taxonomy", required=True, type=Path,
                    help="Path to local clone of oktopi-org/Taxonomy-config")
    ap.add_argument("--plugin", type=Path,
                    default=Path(__file__).resolve().parents[1],
                    help="Path to the plugin root (defaults to the parent of this script's parent)")
    args = ap.parse_args()
    build(args.taxonomy.resolve(), args.plugin.resolve())


if __name__ == "__main__":
    main()
